# -*- coding: utf-8 -*-
"""
Task1 Phase 4.5：bazi_cake.xlsx vs Google Sheets 食材資料比對（純唯讀，不刪檔、不寫入）。

比對範圍：
  - xlsx 的 ingredients tab  vs  Sheet 的 ingredients tab
  - xlsx 的 monthly tab      vs  Sheet 的 monthly tab（若 Sheet 有此 tab）

輸出：兩邊筆數、xlsx 有而 Sheet 沒有的、欄位差異。
xlsx 是否刪除由 Olina 看報告後決定，本腳本不刪、不改任何資料。

執行：python compare_xlsx_vs_sheet.py
"""
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# 後台合併 Sheet ID（Task1 Phase 1 後，與 admin.py BACK_SHEET_ID 一致）
BACK_SHEET_ID = "1oYJ7qO4E40aw1RVip-O3NM6X_U-mYxGT1NIn_YvpW2E"
os.environ.setdefault("GOOGLE_SHEETS_ID", BACK_SHEET_ID)

from openpyxl import load_workbook
import db  # 唯讀走 db._get_sheet


def read_xlsx_tab(tab):
    wb = load_workbook("bazi_cake.xlsx", read_only=True, data_only=True)
    if tab not in wb.sheetnames:
        return None, []
    ws = wb[tab]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    if not rows:
        return [], []
    header = [("" if c is None else str(c).strip()) for c in rows[0]]
    data = []
    for r in rows[1:]:
        cells = [("" if c is None else str(c).strip()) for c in r]
        if any(cells):  # 跳過全空列
            data.append(cells)
    return header, data


def read_sheet_tab(tab):
    """唯讀讀 Sheet 指定 tab；tab 不存在或連線失敗回 (None, [])。"""
    try:
        ws = db._get_sheet(tab)
    except Exception as e:
        print(f"  [警告] 讀取 Sheet tab『{tab}』失敗：{type(e).__name__}: {e}")
        return None, []
    rows = ws.get_all_values()
    if not rows:
        return [], []
    header = [str(c).strip() for c in rows[0]]
    data = []
    for r in rows[1:]:
        cells = [str(c).strip() for c in r]
        if any(cells):
            data.append(cells)
    return header, data


def compare_by_first_col(xlsx_header, xlsx_data, sheet_header, sheet_data, key_label):
    """以第一欄為 key（食材名稱 / 月份）逐筆比對。"""
    print(f"\n  欄位標題：")
    print(f"    xlsx :  {xlsx_header}")
    print(f"    Sheet:  {sheet_header}")
    if xlsx_header != sheet_header:
        print(f"    ⚠️ 標題不一致（欄位數 xlsx={len(xlsx_header)} / Sheet={len(sheet_header)}）")
    else:
        print(f"    ✓ 標題一致")

    xlsx_keys = {row[0]: row for row in xlsx_data if row and row[0]}
    sheet_keys = {row[0]: row for row in sheet_data if row and row[0]}

    print(f"\n  筆數：xlsx={len(xlsx_data)}  Sheet={len(sheet_data)}")

    only_in_xlsx = [k for k in xlsx_keys if k not in sheet_keys]
    only_in_sheet = [k for k in sheet_keys if k not in xlsx_keys]

    print(f"\n  【xlsx 有、Sheet 沒有】共 {len(only_in_xlsx)} 筆：")
    if only_in_xlsx:
        for k in only_in_xlsx:
            print(f"    - {k}：{xlsx_keys[k]}")
    else:
        print(f"    （無）")

    print(f"\n  【Sheet 有、xlsx 沒有】共 {len(only_in_sheet)} 筆：")
    if only_in_sheet:
        for k in only_in_sheet:
            print(f"    - {k}")
    else:
        print(f"    （無）")

    # 兩邊都有的 key，逐欄比對值差異
    common = [k for k in xlsx_keys if k in sheet_keys]
    diffs = []
    for k in common:
        xr = xlsx_keys[k]
        sr = sheet_keys[k]
        n = max(len(xr), len(sr))
        cell_diffs = []
        for i in range(n):
            xv = xr[i] if i < len(xr) else ""
            sv = sr[i] if i < len(sr) else ""
            if xv != sv:
                col = xlsx_header[i] if i < len(xlsx_header) else f"第{i+1}欄"
                cell_diffs.append((col, xv, sv))
        if cell_diffs:
            diffs.append((k, cell_diffs))

    print(f"\n  【兩邊都有但欄位值有差異】共 {len(diffs)} 筆（共同 {len(common)} 筆）：")
    if diffs:
        for k, cell_diffs in diffs:
            print(f"    - {k}：")
            for col, xv, sv in cell_diffs:
                print(f"        {col}: xlsx=「{xv}」 / Sheet=「{sv}」")
    else:
        print(f"    （無，{len(common)} 筆共同資料欄位值完全一致）")

    return {
        "xlsx_count": len(xlsx_data),
        "sheet_count": len(sheet_data),
        "only_in_xlsx": only_in_xlsx,
        "only_in_sheet": only_in_sheet,
        "value_diffs": diffs,
        "header_match": xlsx_header == sheet_header,
    }


def main():
    print("=" * 70)
    print("bazi_cake.xlsx  vs  Google Sheets 食材資料比對報告")
    print(f"後台 Sheet ID：{BACK_SHEET_ID}")
    print("=" * 70)

    summary = {}

    # ── ingredients ──
    print("\n" + "─" * 70)
    print("[1] ingredients tab 比對")
    print("─" * 70)
    xh, xd = read_xlsx_tab("ingredients")
    sh, sd = read_sheet_tab("ingredients")
    if sh is None:
        print("  Sheet 無 ingredients tab 或讀取失敗，略過比對。")
    else:
        summary["ingredients"] = compare_by_first_col(xh, xd, sh, sd, "食材名稱")

    # ── monthly ──
    print("\n" + "─" * 70)
    print("[2] monthly tab 比對")
    print("─" * 70)
    xh, xd = read_xlsx_tab("monthly")
    sh, sd = read_sheet_tab("monthly")
    if sh is None:
        print("  Sheet 無 monthly tab（或讀取失敗）。xlsx 的 monthly 在 Sheet 端無對應。")
        summary["monthly"] = {"sheet_has_tab": False, "xlsx_count": len(xd)}
    else:
        summary["monthly"] = compare_by_first_col(xh, xd, sh, sd, "月份")

    # ── 結論摘要 ──
    print("\n" + "=" * 70)
    print("結論摘要")
    print("=" * 70)
    ing = summary.get("ingredients")
    if ing:
        print(f"ingredients：xlsx {ing['xlsx_count']} 筆 / Sheet {ing['sheet_count']} 筆；"
              f"xlsx 獨有 {len(ing['only_in_xlsx'])} 筆、Sheet 獨有 {len(ing['only_in_sheet'])} 筆、"
              f"值差異 {len(ing['value_diffs'])} 筆。")
    mon = summary.get("monthly")
    if mon and mon.get("sheet_has_tab") is False:
        print(f"monthly：Sheet 端無 monthly tab；xlsx 有 {mon['xlsx_count']} 筆。")
    elif mon:
        print(f"monthly：xlsx {mon['xlsx_count']} 筆 / Sheet {mon['sheet_count']} 筆；"
              f"xlsx 獨有 {len(mon['only_in_xlsx'])} 筆、值差異 {len(mon['value_diffs'])} 筆。")
    print("\n（本腳本純唯讀，未刪除 xlsx、未修改任何 Sheet。是否刪 xlsx 由 Olina 決定。）")


if __name__ == "__main__":
    main()
