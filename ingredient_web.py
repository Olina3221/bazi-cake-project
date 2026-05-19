from flask import Flask, request, jsonify, render_template_string
from ingredient_classifier import classify_ingredient, write_ingredient, EXCEL_PATH
from openpyxl import load_workbook

app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>食材五行判斷</title>
<style>
* { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; background: #f5f5f0; color: #1a1a18; min-height: 100vh; padding: 2rem 1rem; }
.container { max-width: 720px; margin: 0 auto; }
h1 { font-size: 20px; font-weight: 500; margin-bottom: 0.25rem; }
.subtitle { font-size: 14px; color: #888780; margin-bottom: 2rem; }
.card { background: #fff; border: 0.5px solid #ddd; border-radius: 12px; padding: 1.5rem; margin-bottom: 1.5rem; }
.input-row { display: flex; gap: 10px; align-items: center; }
input[type="text"] { flex: 1; height: 40px; border: 0.5px solid #ccc; border-radius: 8px; padding: 0 12px; font-size: 15px; outline: none; }
input[type="text"]:focus { border-color: #378ADD; box-shadow: 0 0 0 3px rgba(55,138,221,0.12); }
button { height: 40px; padding: 0 20px; border-radius: 8px; font-size: 14px; cursor: pointer; border: 0.5px solid #ccc; background: #fff; transition: background 0.15s; white-space: nowrap; }
button:hover { background: #f1efe8; }
button.primary { background: #1D9E75; color: #fff; border-color: #1D9E75; }
button.primary:hover { background: #0F6E56; }
button:disabled { opacity: 0.4; cursor: not-allowed; }
.hint { font-size: 13px; color: #888780; margin-top: 8px; }
.result-card { background: #fff; border: 0.5px solid #ddd; border-radius: 12px; padding: 1.5rem; margin-bottom: 1rem; }
.wuxing-badge { display: inline-block; padding: 2px 12px; border-radius: 20px; font-size: 13px; font-weight: 500; margin-right: 6px; }
.wood  { background: #EAF3DE; color: #27500A; }
.fire  { background: #FAECE7; color: #712B13; }
.earth { background: #FAEEDA; color: #633806; }
.metal { background: #F1EFE8; color: #444441; }
.water { background: #E6F1FB; color: #0C447C; }
.grid { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 10px; margin: 1rem 0; }
.grid-item { background: #f5f5f0; border-radius: 8px; padding: 10px 12px; }
.grid-label { font-size: 11px; color: #888780; margin-bottom: 4px; }
.grid-value { font-size: 14px; color: #1a1a18; }
.action-row { display: flex; gap: 10px; margin-top: 1rem; }
.status { padding: 10px 14px; border-radius: 8px; font-size: 14px; margin-bottom: 1rem; }
.status.success { background: #EAF3DE; color: #27500A; }
.status.error   { background: #FCEBEB; color: #791F1F; }
.status.warning { background: #FAEEDA; color: #633806; }
.loading { display: flex; align-items: center; gap: 10px; color: #888780; font-size: 14px; padding: 1rem 0; }
.spinner { width: 18px; height: 18px; border: 2px solid #ddd; border-top-color: #1D9E75; border-radius: 50%; animation: spin 0.7s linear infinite; }
@keyframes spin { to { transform: rotate(360deg); } }
table { width: 100%; border-collapse: collapse; font-size: 13px; margin-top: 1rem; }
th { text-align: left; padding: 8px 10px; background: #f5f5f0; font-weight: 500; color: #888780; border-bottom: 0.5px solid #ddd; }
td { padding: 8px 10px; border-bottom: 0.5px solid #f1efe8; color: #1a1a18; }
tr:last-child td { border-bottom: none; }
.section-title { font-size: 15px; font-weight: 500; margin-bottom: 1rem; }
</style>
</head>
<body>
<div class="container">
  <h1>食材五行判斷</h1>
  <p class="subtitle">輸入食材名稱，由 Claude 分析五行屬性並寫入資料庫</p>

  <div class="card">
    <div class="input-row">
      <input type="text" id="ingredient-input" placeholder="例：百香果、椰子、紫薯" autofocus>
      <button class="primary" onclick="classify()" id="classify-btn">判斷五行</button>
    </div>
    <p class="hint">可輸入多個食材，用逗號分隔。例：百香果, 椰子, 紫薯</p>
  </div>

  <div id="status-area"></div>
  <div id="result-area"></div>

  <div class="card">
    <p class="section-title">現有食材資料庫</p>
    <table id="db-table">
      <thead><tr><th>食材</th><th>主五行</th><th>副五行</th><th>用途</th><th>季節</th></tr></thead>
      <tbody id="db-body"></tbody>
    </table>
  </div>
</div>

<script>
const wuxingClass = { "木":"wood","火":"fire","土":"earth","金":"metal","水":"water" };
const wuxingLabel = { "木":"木 🌿","火":"火 🔥","土":"土 🏔","金":"金 ✨","水":"水 💧" };

function badge(w) {
  if (!w) return '<span style="color:#aaa">—</span>';
  return `<span class="wuxing-badge ${wuxingClass[w]||''}">${wuxingLabel[w]||w}</span>`;
}

async function classify() {
  const raw = document.getElementById("ingredient-input").value.trim();
  if (!raw) return;
  const names = raw.split(/[,，、]/).map(s=>s.trim()).filter(Boolean);
  const btn = document.getElementById("classify-btn");
  btn.disabled = true;
  document.getElementById("status-area").innerHTML = "";
  document.getElementById("result-area").innerHTML = "";

  for (const name of names) {
    showLoading(`正在分析：${name}...`);
    try {
      const res = await fetch("/classify", {
        method: "POST",
        headers: {"Content-Type":"application/json"},
        body: JSON.stringify({name})
      });
      const data = await res.json();
      if (data.error) { showStatus("error", `錯誤：${data.error}`); continue; }
      renderResult(data);
    } catch(e) {
      showStatus("error", `連線失敗：${e.message}`);
    }
  }
  document.getElementById("status-area").innerHTML = "";
  btn.disabled = false;
  loadDB();
}

function showLoading(msg) {
  document.getElementById("status-area").innerHTML =
    `<div class="loading"><div class="spinner"></div>${msg}</div>`;
}

function showStatus(type, msg) {
  document.getElementById("status-area").innerHTML =
    `<div class="status ${type}">${msg}</div>`;
}

function renderResult(d) {
  const area = document.getElementById("result-area");
  const div = document.createElement("div");
  div.className = "result-card";
  div.id = "result-" + d.食材名稱;
  div.innerHTML = `
    <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;">
      <span style="font-size:16px;font-weight:500;">${d.食材名稱}</span>
      ${badge(d.主五行)}${d.副五行 ? badge(d.副五行) : ""}
    </div>
    <div class="grid">
      <div class="grid-item"><div class="grid-label">用途</div><div class="grid-value">${d.用途||"—"}</div></div>
      <div class="grid-item"><div class="grid-label">季節</div><div class="grid-value">${d.季節||"—"}</div></div>
      <div class="grid-item"><div class="grid-label">備注</div><div class="grid-value">${d.備注||"—"}</div></div>
      <div class="grid-item"><div class="grid-label">顏色分析</div><div class="grid-value">${d.顏色分析||"—"}</div></div>
      <div class="grid-item"><div class="grid-label">氣味分析</div><div class="grid-value">${d.氣味分析||"—"}</div></div>
      <div class="grid-item"><div class="grid-label">屬性分析</div><div class="grid-value">${d.屬性分析||"—"}</div></div>
    </div>
    <div class="action-row">
      <button class="primary" onclick="save('${d.食材名稱}')">寫入資料庫</button>
      <button onclick="discard('${d.食材名稱}')">略過</button>
    </div>
    <div id="save-status-${d.食材名稱}" style="margin-top:8px;"></div>
  `;
  area.appendChild(div);
}

async function save(name) {
  const res = await fetch("/save", {
    method:"POST",
    headers:{"Content-Type":"application/json"},
    body: JSON.stringify({name})
  });
  const data = await res.json();
  const el = document.getElementById("save-status-" + name);
  if (data.status === "saved") {
    el.innerHTML = '<span style="color:#1D9E75;font-size:13px;">✓ 已寫入 Excel</span>';
    document.querySelector(`#result-${name} .action-row`).innerHTML = "";
    loadDB();
  } else if (data.status === "already_exists") {
    el.innerHTML = '<span style="color:#BA7517;font-size:13px;">已存在，略過</span>';
  } else {
    el.innerHTML = `<span style="color:#E24B4A;font-size:13px;">錯誤：${data.error}</span>`;
  }
}

function discard(name) {
  const el = document.getElementById("result-" + name);
  if (el) el.remove();
}

async function loadDB() {
  const res = await fetch("/ingredients");
  const data = await res.json();
  const tbody = document.getElementById("db-body");
  tbody.innerHTML = data.map(r => `
    <tr>
      <td><strong>${r[0]}</strong></td>
      <td>${badge(r[1])}</td>
      <td>${r[2] ? badge(r[2]) : '<span style="color:#aaa">—</span>'}</td>
      <td style="color:#888780;font-size:12px;">${r[3]||"—"}</td>
      <td style="color:#888780;font-size:12px;">${r[4]||"—"}</td>
    </tr>`).join("");
}

document.getElementById("ingredient-input").addEventListener("keydown", e => {
  if (e.key === "Enter") classify();
});

loadDB();
</script>
</body>
</html>"""

@app.route("/")
def index():
    return render_template_string(HTML)

@app.route("/classify", methods=["POST"])
def classify():
    name = request.json.get("name", "").strip()
    if not name:
        return jsonify({"error": "請輸入食材名稱"})
    try:
        result = classify_ingredient(name)
        app.config["pending"][name] = result
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/save", methods=["POST"])
def save():
    name = request.json.get("name", "").strip()
    result = app.config["pending"].get(name)
    if not result:
        return jsonify({"error": "找不到判斷結果，請重新判斷"})
    try:
        ok, status = write_ingredient(result)
        return jsonify({"status": status})
    except Exception as e:
        return jsonify({"error": str(e)})

@app.route("/ingredients")
def ingredients():
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["ingredients"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            rows.append(list(row[:5]))
    return jsonify(rows)

if __name__ == "__main__":
    app.config["pending"] = {}
    print("🎂 食材五行判斷工具啟動中...")
    print("👉 開啟瀏覽器前往：http://localhost:5001")
    app.run(debug=False, port=5001)
