import anthropic
import json
from openpyxl import load_workbook

EXCEL_PATH = "bazi_cake.xlsx"

SYSTEM_PROMPT = """你是一位有二十年實戰經驗的八字命理師，同時也是資深烘焙食材研究者。
你深諳五行學理，並將其應用在蛋糕食材的能量設計上。你的判斷被台灣多位知名八字蛋糕師傅視為業界標準。

## 你的判斷哲學

五行歸屬不是機械配對，而是理解食材的「能量本質」：
- 食材的**天然本質**（原始植物/動物屬性）優先於加工形態
- **複合食材**（如「藍莓庫利凍」）以主食材為核心，加工方式為輔助修正
- 加工形態的影響：凍/果凍（稍帶金的收斂，但不改變主五行）、醬（加熱濃縮，略帶土）、泥（土性更強）、乾燥（金性增強）
- 當食材同時具備多種五行特質時，以「在蛋糕中最主要呈現的能量感」為準

## 五行對應體系（資深判斷版）

### 木🌿
- 顏色：翠綠、嫩綠、青色、黃綠
- 氣味：清新草香、茶香、青草、微酸爽
- 屬性：生發向上、酸涼清爽、疏肝理氣
- 代表食材：抹茶、薄荷、檸檬、奇異果、青蘋果、百香果（酸爽為主）

### 火🔥
- 顏色：鮮紅、桃紅、珊瑚橘、玫瑰粉
- 氣味：甜香花香、果香濃郁、溫暖馥郁
- 屬性：溫熱活化、促進循環、開心醒神
- 代表食材：草莓、覆盆子、荔枝、玫瑰、紅心火龍果、櫻桃

### 土🏔️
- 顏色：米黃、棕色、大地色、橙黃
- 氣味：甜膩奶香、焦糖、麥芽、樸實甘甜
- 屬性：滋養補中、厚重紮實、安神穩定
- 代表食材：芋頭、焦糖、南瓜、芒果、栗子、肉桂、香草（奶香重的）

### 金✨
- 顏色：純白、乳白、金色、銀色、透明
- 氣味：清淡高雅、辛香收斂、輕盈潔淨
- 屬性：收斂精緻、清肺潤燥、質感高雅
- 代表食材：香草籽奶油、白桃、水梨、奶酪、食用金箔、迷迭香

### 水💧
- 顏色：深藍、深紫、黑色、深色系
- 氣味：深邃沉靜、微苦、無味或淡味
- 屬性：涼性滋潤、沉降收藏、補腎深養
- 代表食材：藍莓、紫葡萄、黑芝麻、紫薯、藍莓系製品

## 複合食材判斷範例（你的判斷標準）

- 「藍莓庫利凍」→ 主食材藍莓（水），庫利凍是打碎過濾後的濃縮果汁凍，保留藍莓深紫色澤與酸甜滋味，主五行=水，副五行=木（酸味）
- 「草莓庫利凍」→ 草莓（火），庫利凍形態，主五行=火，副五行=木
- 「焦糖奶油」→ 焦糖（土），奶油（金/土），主五行=土，副五行=金
- 「香草籽鮮奶油」→ 香草（金/土）+鮮奶油（金），主五行=金
- 「黑芝麻鮮奶油」→ 黑芝麻（水）+鮮奶油（金），主五行=水，副五行=金
- 「玫瑰鮮奶油」→ 玫瑰（火）+鮮奶油（金），主五行=火，副五行=金

## 輸出格式

判斷完成後，只回傳以下 JSON，不要加任何說明或 markdown 包裝：
{
  "食材名稱": "...",
  "主五行": "木/火/土/金/水",
  "副五行": "木/火/土/金/水 或空字串",
  "用途": "底色/夾層/秀面/點綴（可複選，逗號分隔）",
  "季節": "常備/春/夏/秋/冬（可複選，逗號分隔）",
  "顏色分析": "一句話說明顏色對應",
  "氣味分析": "一句話說明氣味對應",
  "屬性分析": "一句話說明屬性對應",
  "備注": "有特殊說明才填，否則空字串"
}"""


def classify_ingredient(name: str) -> dict:
    client = anthropic.Anthropic()
    message = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=600,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": f"請判斷食材：{name}"}]
    )
    raw = message.content[0].text.strip()
    raw = raw.replace("```json", "").replace("```", "").strip()
    return json.loads(raw)


def ingredient_exists(ws, name: str) -> bool:
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == name:
            return True
    return False


def write_ingredient(result: dict):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["ingredients"]

    if ingredient_exists(ws, result["食材名稱"]):
        wb.close()
        return False, "already_exists"

    ws.append([
        result.get("食材名稱", ""),
        result.get("主五行", ""),
        result.get("副五行", ""),
        result.get("用途", ""),
        result.get("季節", ""),
        result.get("顏色分析", ""),
        result.get("氣味分析", ""),
        result.get("屬性分析", ""),
        result.get("備注", ""),
    ])
    wb.save(EXCEL_PATH)
    return True, "saved"


def add_ingredient(name: str):
    print(f"\n🔍 判斷中：{name}")
    result = classify_ingredient(name)

    print(f"\n📋 判斷結果：")
    print(f"   食材：{result['食材名稱']}")
    print(f"   主五行：{result['主五行']}  副五行：{result.get('副五行') or '無'}")
    print(f"   用途：{result['用途']}")
    print(f"   季節：{result['季節']}")
    print(f"   顏色：{result['顏色分析']}")
    print(f"   氣味：{result['氣味分析']}")
    print(f"   屬性：{result['屬性分析']}")
    if result.get("備注"):
        print(f"   備注：{result['備注']}")

    confirm = input("\n✅ 確認寫入 Excel？(y/n/修改): ").strip().lower()

    if confirm == "y":
        ok, status = write_ingredient(result)
        if ok:
            print(f"✅ 已寫入：{name}")
        elif status == "already_exists":
            print(f"⚠️  {name} 已存在，略過")
    elif confirm == "n":
        print("❌ 取消")
    else:
        print("（如需修改請直接在 Excel 中調整）")

    return result


def batch_add(names: list):
    for name in names:
        add_ingredient(name.strip())


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        ingredients = sys.argv[1:]
        if len(ingredients) == 1:
            add_ingredient(ingredients[0])
        else:
            for name in ingredients:
                add_ingredient(name)
    else:
        print("八字蛋糕食材五行判斷工具")
        print("用法：")
        print("  python ingredient_classifier.py 百香果")
        print("  python ingredient_classifier.py 百香果 椰子 紫薯")
        print("\n或在 Python 中 import：")
        print("  from ingredient_classifier import add_ingredient")
        print("  add_ingredient('百香果')")
