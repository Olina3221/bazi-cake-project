from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_workbook():
    wb = Workbook()

    # ── 1. orders 工作表 ──────────────────────────────────────
    ws_orders = wb.active
    ws_orders.title = "orders"

    orders_headers = [
        "訂單編號", "建立日期", "取件日期",
        "客人姓名", "電話", "生日", "出生時辰", "出生地",
        "心願大方向", "心願細分", "備註",
        "底色", "第一夾層", "第二夾層", "秀面", "點綴",
        "文案", "狀態"
    ]
    ws_orders.append(orders_headers)

    header_fill = PatternFill("solid", start_color="4A90D9")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col, _ in enumerate(orders_headers, 1):
        cell = ws_orders.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 12, 12, 10, 14, 12, 10, 10, 12, 12, 20, 12, 12, 12, 12, 12, 40, 10]
    for i, w in enumerate(col_widths, 1):
        ws_orders.column_dimensions[get_column_letter(i)].width = w

    # ── 2. ingredients 工作表 ──────────────────────────────────
    ws_ing = wb.create_sheet("ingredients")

    ing_headers = ["食材名稱", "主五行", "副五行", "用途", "季節", "顏色分析", "氣味分析", "屬性分析", "備注"]
    ws_ing.append(ing_headers)

    ing_fill = PatternFill("solid", start_color="5DAD7A")
    ing_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col, _ in enumerate(ing_headers, 1):
        cell = ws_ing.cell(row=1, column=col)
        cell.fill = ing_fill
        cell.font = ing_font
        cell.alignment = Alignment(horizontal="center")

    ing_widths = [14, 10, 10, 20, 20, 16, 16, 16, 20]
    for i, w in enumerate(ing_widths, 1):
        ws_ing.column_dimensions[get_column_letter(i)].width = w

    # 預填初始食材
    initial_ingredients = [
        ("薄荷葉",   "木", "",   "秀面,點綴", "常備",       "鮮綠",   "清涼",   "涼性"),
        ("迷迭香",   "金", "木", "秀面,點綴", "常備",       "銀灰綠", "辛香",   "溫性"),
        ("香草籽",   "金", "土", "夾層,秀面", "常備",       "深褐",   "溫甜",   "溫性"),
        ("抹茶碎",   "木", "",   "秀面",      "常備",       "深綠",   "清苦",   "涼性"),
        ("檸檬皮絲", "木", "土", "秀面,點綴", "常備",       "黃綠",   "清酸",   "涼性"),
        ("草莓",     "火", "",   "夾層,秀面", "春,冬",      "紅色",   "甜香",   "溫性"),
        ("覆盆子",   "火", "",   "夾層,秀面", "常備",       "深紅",   "酸甜",   "涼性"),
        ("藍莓",     "水", "",   "夾層,秀面", "常備",       "深藍紫", "微酸",   "涼性"),
        ("食用金箔", "金", "",   "點綴",      "常備",       "金色",   "無味",   "加工食材"),
        ("銀珠糖",   "金", "水", "點綴",      "常備",       "銀白",   "甜",     "加工食材"),
        ("焦糖脆片", "土", "火", "秀面",      "常備",       "棕色",   "焦甜",   "溫性"),
        ("黑芝麻",   "水", "",   "夾層",      "常備",       "黑色",   "香濃",   "溫性"),
        ("芋頭泥",   "土", "水", "夾層",      "常備",       "紫灰",   "甜香",   "平性"),
        ("玫瑰花瓣", "火", "木", "點綴",      "常備",       "粉紅",   "花香",   "溫性"),
    ]
    for row in initial_ingredients:
        ws_ing.append(row)

    # ── 3. monthly 工作表 ─────────────────────────────────────
    ws_monthly = wb.create_sheet("monthly")

    monthly_headers = ["月份", "主力五行", "第二五行", "建議備料食材", "當月訂單數"]
    ws_monthly.append(monthly_headers)

    mo_fill = PatternFill("solid", start_color="9B6BB5")
    mo_font = Font(bold=True, color="FFFFFF", name="Arial")
    for col, _ in enumerate(monthly_headers, 1):
        cell = ws_monthly.cell(row=1, column=col)
        cell.fill = mo_fill
        cell.font = mo_font
        cell.alignment = Alignment(horizontal="center")

    monthly_data = [
        (1,  "金", "水", "藍莓、水蜜桃罐頭、蘋果、銀珠糖"),
        (2,  "火", "金", "草莓、食用金箔、香草籽"),
        (3,  "木", "火", "薄荷葉、草莓、檸檬皮絲"),
        (4,  "木", "水", "薄荷葉、藍莓、桑椹"),
        (5,  "火", "土", "草莓、芒果、焦糖脆片"),
        (6,  "水", "火", "藍莓、紅心火龍果、覆盆子"),
        (7,  "水", "木", "藍莓、薄荷、奇異果"),
        (8,  "土", "火", "芒果、哈密瓜、荔枝"),
        (9,  "土", "金", "鳳梨、水梨、焦糖脆片"),
        (10, "金", "土", "梨子、蘋果、迷迭香"),
        (11, "水", "金", "藍莓、紫葡萄、食用金箔"),
        (12, "水", "火", "藍莓、草莓、紫薯"),
    ]
    for row in monthly_data:
        ws_monthly.append(row)

    mo_widths = [8, 10, 10, 40, 12]
    for i, w in enumerate(mo_widths, 1):
        ws_monthly.column_dimensions[get_column_letter(i)].width = w

    wb.save("bazi_cake.xlsx")
    print("✅ bazi_cake.xlsx 建立完成")
    print(f"   - orders：{len(orders_headers)} 個欄位")
    print(f"   - ingredients：{len(initial_ingredients)} 筆初始食材")
    print(f"   - monthly：12 個月份備料建議")

if __name__ == "__main__":
    create_workbook()
